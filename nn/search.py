import torch
from langchain_core.embeddings import Embeddings
from langchain_community.retrievers import BM25Retriever
from langchain.retrievers import EnsembleRetriever
from nltk.stem.snowball import SnowballStemmer
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import re
from tqdm import tqdm
import json #added for key_words

from openpyxl import load_workbook #added for keywords
from collections import defaultdict #added for keywords


_STEMMER = SnowballStemmer("russian")
_PREPROCESS_REGEX = re.compile(r'[^а-яё\s]')  
_STOP_WORDS = set(stopwords.words('russian'))
_BANNED_WORDS = {'мгу', 'физфак', 'физический', 'университет'}
_STEMMED_BANNED_WORDS = {_STEMMER.stem(w) for w in _BANNED_WORDS}

def preprocess(text):
    cleaned = _PREPROCESS_REGEX.sub('', text.lower())
    words = word_tokenize(cleaned, language="russian")
    filtered_tokens = [word for word in words if word.strip() and word not in _STOP_WORDS]
    stemmed_words = [_STEMMER.stem(word) for word in filtered_tokens]
    return [word for word in stemmed_words if word not in _STEMMED_BANNED_WORDS]

class E5LangChainEmbedder(Embeddings):
    def __init__(
        self,
        tokenizer,
        model,
        device: str = 'cpu',
        embed_batch_size: int = 8,
        add_prefix: bool = False,
        disable_tqdm: bool = False
    ):
        self.tokenizer = tokenizer
        self.model = model.to(device)
        self.device = device
        self.embed_batch_size = embed_batch_size
        self.add_prefix = add_prefix
        self.disable_tqdm = disable_tqdm
        self.model.eval()

    def _average_pool(self, last_hidden_states, attention_mask):
        last_hidden = last_hidden_states.masked_fill(~attention_mask[..., None].bool(), 0.0)
        return last_hidden.sum(dim=1) / attention_mask.sum(dim=1)[..., None]

    def embed_documents(self, texts):
        if self.add_prefix:
            texts = ["passage: " + t for t in texts]

        all_embeddings = []
        for i in tqdm(range(0, len(texts), self.embed_batch_size),
                     desc="Вычисление эмбеддингов", unit="batch",
                     disable=self.disable_tqdm):
            batch_texts = texts[i:i + self.embed_batch_size]
            batch_dict = self.tokenizer(
                batch_texts,
                max_length=512,
                padding=True,
                truncation=True,
                return_tensors='pt'
            ).to(self.device)

            with torch.no_grad():
                outputs = self.model(**batch_dict)
                embeddings = self._average_pool(
                    outputs.last_hidden_state,
                    batch_dict['attention_mask']
                )
                embeddings = torch.nn.functional.normalize(embeddings, p=2, dim=1)
                all_embeddings.extend(embeddings.cpu().tolist())

        return all_embeddings

    def embed_query(self, text):
        if self.add_prefix:
            text = "query: " + text
        
        batch_dict = self.tokenizer(
            [text],
            max_length=512,
            padding=True,
            truncation=True,
            return_tensors='pt'
        ).to(self.device)
        
        with torch.no_grad():
            outputs = self.model(**batch_dict)
            embeddings = self._average_pool(
                outputs.last_hidden_state,
                batch_dict['attention_mask']
            )
            embeddings = torch.nn.functional.normalize(embeddings, p=2, dim=1)
            return embeddings.cpu().tolist()[0]
    
    
# def get_context(query, tokenizer, model, bm_25, vector_store, ensemble_k=5, retrivier_k=10):
#     bm_25.k = retrivier_k
#     vector_retriever = vector_store.as_retriever(search_kwargs={"k": retrivier_k})

#     ensemble_retriever = EnsembleRetriever(
#         retrievers=[bm_25, vector_retriever],
#         weights=[0.5, 0.5]
#     )

#     raiting = ensemble_retriever.invoke(query)[:ensemble_k]

#     results = []
#     for res in raiting:
#         results.append({
#         "topic": res.metadata['source'],
#         "full_text": res.page_content
#     })

#     combined_text = "\n".join(doc.page_content for doc in raiting)
    
#     return results, combined_text


#---added for keywords---
def generate_keywords_dict(excel_path, output_json_path=None):
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        keywords_dict = defaultdict(list)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:
                continue

            topic_id = str(row[3]).strip()
            key_words_value = row[4]

            if key_words_value is None or not str(key_words_value).strip():
                continue

            keywords = [
                kw.strip().lower()
                for kw in str(key_words_value).split(",")
                if kw.strip()
            ]

            for kw in keywords:
                keywords_dict[kw].append(topic_id)

        result = dict(keywords_dict)

        # Выводим, сколько ключевых слов найдено
        print(f"Dictionary created: {len(result)} keywords")

        if output_json_path:
            with open(output_json_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            print(f"Dictionary saved to {output_json_path}")

        return result

    except Exception as e:
        print(f"Error: {e}")
        return {}

# def test_func(text):
#      'фунция для примера'
#      dict_of_keys = {'Профком': [1, 7, 89]}

#      if len(text) < 3 and any(text) in dict_of_keys.keys():
#           context = key_words_search()
#      else:
#          context = semantinc_search()

def key_words_search(query, key_words_dict, vector_store, ensemble_k=5, verbose=False):
    words = query.lower().split()
    matching_ids = set()
    
    # Собираем ID тем, соответствующих ключевым словам
    for word in words:
        if word in key_words_dict:
            matching_ids.update(int(excel_id) for excel_id in key_words_dict[word])
        
    if verbose:
        print(f"Key words search: Found {len(matching_ids)} matching documents")
    
    # Создаём фильтр по ID
    filter_criteria = {"id": {"$in": list(matching_ids)}} 
    
    docs = vector_store.as_retriever(
        search_kwargs={"filter": filter_criteria, "k": len(matching_ids)}
    ).get_relevant_documents("")
    
    # Форматируем результат
    results = [
        {"topic": doc.metadata["source"], "full_text": doc.page_content}
        for doc in docs[:ensemble_k]
    ]
    combined_text = "\n".join(doc.page_content for doc in docs)
    
    return results, combined_text


def semantic_search(query, bm_25, vector_store, ensemble_k=5, retriever_k=10, verbose=False):
    if verbose:
        print("Semantic search: Using hybrid retrieval (BM25 + vector search)")
    
    bm_25.k = retriever_k
    vector_retriever = vector_store.as_retriever(search_kwargs={"k": retriever_k})

    ensemble_retriever = EnsembleRetriever(
        retrievers=[bm_25, vector_retriever],
        weights=[0.5, 0.5]
    )

    rankings = ensemble_retriever.invoke(query)[:ensemble_k]

    results = []
    for res in rankings:
        results.append({
            "topic": res.metadata['source'],
            "full_text": res.page_content
        })

    combined_text = "\n".join(doc.page_content for doc in rankings)
    
    return results, combined_text


def get_context(query, key_words_dict, bm_25, vector_store, 
                ensemble_k=5, retriever_k=10, verbose=True):
    words = query.lower().split()
    
    if len(words) < 3 and any(word in key_words_dict for word in words):
        if verbose:
            print("→ Using KEY WORDS SEARCH")
        return key_words_search(query, key_words_dict, vector_store, ensemble_k, verbose)
    
    if verbose:
        print("→ Using SEMANTIC SEARCH")
    return semantic_search(query, bm_25, vector_store, ensemble_k, retriever_k, verbose)